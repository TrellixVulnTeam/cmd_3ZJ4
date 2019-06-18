#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""

My very usefull tools library
require pandas and dask!!

MIT License

"""
__version__ = "0.2.0"
__author__ = "m.yama"

__all__ = [
        "dd",
        "read_csv",
        "read_excel",
        "read_json",
        "read_any",
        "df_cast",
        "dflines",
        "vdf",
        "hdf",
        ]


from util.core import (
        BUF,
        vmfree,
        isdataframe,
        Path,
        TMPDIR
        )

import re
import sys
from io import TextIOBase, BufferedIOBase, StringIO, BytesIO

# 3rd party modules
import numpy as np

import pandas.io.formats.excel
header_style = {"font": {"bold": True},
                "borders": {"top": "thin",
                            "right": "thin",
                            "bottom": "thin",
                            "left": "thin"},
                "alignment": {"horizontal": "center",
                              "vertical": "center"}}
pandas.io.formats.excel.ExcelFormatter.header_style = header_style

import pandas.core.generic
def to_excel_bordered(self, excel_writer, sheet_name="Sheet1", na_rep="",
             float_format=None, columns=None, header=True, index=True,
             index_label=None, startrow=0, startcol=0, engine="openpyxl",
             merge_cells=True, encoding=None, inf_rep="inf", verbose=True,
             freeze_panes=None):


    kw = locals()
    if "self" in kw:
        del kw["self"]

    from openpyxl.utils import get_column_letter

    writer = excel_writer if hasattr(excel_writer, "book") else pd.ExcelWriter(excel_writer, engine=engine)
    kw["excel_writer"] = writer
    prop = {"border-style": "solid", "border-width": "thin", "vertical-align": "center"}
    self.style.set_properties(**prop).to_excel(**kw)

    sheet = writer.sheets["Sheet1"]

    # auto fit column width
    startrow += 2 if header else 1
    startcol += 2 if index else 1

    for i, (col, s) in enumerate(self.astype(str).items(), startcol):
        width = min([max([s.str.len().max(), len(str(col)), 8]) + 1, 130])
        col = get_column_letter(i)
        sheet.column_dimensions[col].width = width

    if not hasattr(excel_writer, "book"):
        writer.save()
        writer.close()
pandas.core.generic.NDFrame.to_excel_bordered = to_excel_bordered

def to_excel_plus(self, excel_writer, sheet_name="Sheet1",
             conditional_value=None, autofilter=True, title=None,
             na_rep="",
             float_format=None, columns=None, header=True, index=True,
             index_label=None, startrow=0, startcol=0, engine="openpyxl",
             merge_cells=True, encoding=None, inf_rep="inf", verbose=True,
             freeze_panes=None):


    if title:
        startrow += 2

    kw = locals()
    if "self" in kw:
        del kw["self"]
        del kw["conditional_value"]
        del kw["autofilter"]
        del kw["title"]

    from openpyxl.formatting import Rule
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import differential, Font, PatternFill

    writer = excel_writer if hasattr(excel_writer, "book") else pd.ExcelWriter(excel_writer, engine=engine)
    kw["excel_writer"] = writer

    self.to_excel_bordered(**kw)

    sheet = writer.sheets["Sheet1"]

    if title:
        sheet["A1"] = title
        sheet['A1'].font = Font(size=16, bold=True)
    if conditional_value:
        dxf = differential.DifferentialStyle(
            font=Font(color="9C0006"),
            fill=PatternFill(bgColor="FFC7CE"))


        startrow += 2 if header else 1
        if index:
            startcol += 1
        st = sheet.cell(row=startrow, column=startcol).coordinate
        en = sheet.cell(row=self.__len__()+startrow, column=len(self.columns)+startcol).coordinate
        rule = Rule(
            type="containsText", operator="containsText",
            formula = ['NOT(ISERROR(SEARCH("{}",{})))'.format(conditional_value, st)],
            dxf=dxf)

        sheet.conditional_formatting.add(st+":"+en, rule)

    if autofilter:
        if not conditional_value and title:
            startrow += 2 if header else 1
        shape = (get_column_letter(sheet.min_column), startrow-1,
                 get_column_letter(sheet.max_column), self.__len__()+startrow-1)
        sheet.auto_filter.ref = "{}{}:{}{}".format(*shape)

    if not hasattr(excel_writer, "book"):
        writer.save()
        writer.close()
pandas.core.generic.NDFrame.to_excel_plus = to_excel_plus

import pandas as pd
import dask
import pathlib
def read_bytes(urlpath, *args, **kw):
    if isinstance(urlpath, pathlib.Path):
        if hasattr(urlpath, "is_compress") and urlpath.is_compress():
            urlpath = urlpath.open().extract(path=TMPDIR)
        else:
            urlpath = str(urlpath)
    elif isinstance(urlpath, (TextIOBase, BufferedIOBase)):
        if hasattr(urlpath, "name"):
            urlpath = urlpath.name
        elif isinstance(urlpath, StringIO):
            pth = TMPDIR + "/StringIO.dat"
            with open(pth, "w") as f:
                f.write(urlpath.getvalue())
            urlpath = pth
        elif isinstance(urlpath, BytesIO):
            pth = TMPDIR + "/BytesIO.dat"
            with open(pth, "wb") as f:
                f.write(urlpath.getvalue())
            urlpath = pth
    elif hasattr(urlpath, "extract"):
        urlpath = urlpath.extract(path=TMPDIR)

    return dask.bytes.core.read_bytes(urlpath, *args, **kw)

#for dask overwride
def sort_values(self, *args, **kw):
    def xsortv(x):
        return x.sort_values(*args, **kw)

    if any(x in kw for x in ["inplace", "axis"]):
        raise AttributeError("not support arguments")
    return self.map_partitions(xsortv)

def nunique(self, dropna=True):
    def nuniq(x):
        return x.nunique(dropna=dropna)
    return self.map_partitions(nuniq)

import dask.dataframe
dask.dataframe.io.csv.read_bytes = read_bytes
dask.dataframe.core._Frame.sort_values = sort_values
dask.dataframe.core._Frame.nunique = nunique
import dask.dataframe as dd

try:
    from xlrd import XLRDError
except ImportError:
    XLRDError = ImportError


forcedask = False

class _dfhandler(object):

    def __init__(self, func, path_or_buffer, *args, **kw):
        self.func = func
        self.path_or_buffer = Path(path_or_buffer)
        self.args = args
        self.kw = kw
        self.concater = None
        self.gk = []
        self.size = 0
        #TODO UnicoDedecodeError => zopen_recursive
        self.guesskw()

    def guesskw(self):
        def kwargs(p):
            #TODO botolneck 200msec
            return dict(
                sep          = p.sep,
                encoding= p.encoding,
                quoting   = p.quoting,
                delimiter=p.delimiter,
                doublequote= p.doublequote,
                quotechar=p.quotechar,
                )

        p = self.path_or_buffer

        if p.is_compress():
            z = p.open()
            self.size += p.getsize()
            self.gk.append([z, kwargs(z)])
        else:
            self.size += p.getsize()
            self.gk.append([p, kwargs(p)])

    def _docargs(self, obj, re_sp=re.compile("Parameters|Returns"),re_attr = re.compile("\n([^->\s\*][^-\s]*)\s.*:[^\n]+")):
        return set(re_attr.findall(re_sp.split(obj.__doc__)[1]))

    def _handler_read(self):
        #TODO botolneck 142msec
        if forcedask is False and (self.size < BUF or vmfree() > self.size * 5):
            reader = pd.__getattribute__(self.func)
            self.concater = pd.concat
            funckw = self._docargs(pd.read_csv)

        else:
            reader = dd.__getattribute__(self.func) #TODO if dask uri path need
            self.concater = dd.concat
            self.kw["blocksize"] = None
            funckw = self._docargs(reader) | self._docargs(pd.read_csv)
            funckw.remove("filepath_or_buffer")

        for fn, gkw in self.gk:
            gkw.update(self.kw)
            gkw = {x: gkw[x] for x in set(gkw) & set(funckw)}
            yield reader(fn, *self.args, **gkw)

    def compute(self, concatenate=True):
        ret = self._handler_read()
        if len(self.gk) == 1:
            return next(ret)
        if concatenate:
            return self.concater(ret)
        else:
            return list(ret)

def read_csv(f, dtype="object", keep_default_na=False, concatenate=True, *args, **kw):
    def reader(*args, **kw):
        dfh = _dfhandler("read_csv", f, dtype=dtype, low_memory=False,
                            keep_default_na=keep_default_na, *args, **kw)
        return dfh.compute(concatenate)

    try:
        return reader(*args, **kw)
    except pd.errors.ParserError:
        try:
            return reader(sep=",", *args, **kw)
        except pd.errors.ParserError:
            return reader(sep="\t", *args, **kw)


def read_excel(f, sheet_name=None, dtype=str, keep_default_na=False, *arg, **kw):
    try:
        df = pd.read_excel(f,
                             dtype=dtype,
                             keep_default_na=keep_default_na,
                             sheet_name=sheet_name,
                             *arg, **kw)
    except XLRDError:
        raise XLRDError("Unknown Excel Password")
    else:
        if isinstance(df, dict):
            return {k: df_cast(d) for k, d in df.items()}
        else:
            return df_cast(df)

def render_excel(df, outfile="Book1.xlsx", sheetname="Sheet1", skiprows=0,
    conditional_value=None, autofilter=True, title=None):

    from xlsxwriter import Workbook

    with Workbook(outfile) as book:
        sheet = book.add_worksheet(sheetname)
        if title:
            title_fmt = book.add_format(
                dict(border=True, align="left", bold=True, font_size=20))
            sheet.write(skiprows, 0, title, title_fmt)
            skiprows += 1

        header_fmt = book.add_format(
            dict(border=True, align="center", bold=True))

        border_fmt = book.add_format(
            dict(border=True))



        for i, (col, s) in enumerate(df.items()):
            sheet.set_column(i, i, max([s.str.len().max(), len(col)]) + 1)
            sheet.write(skiprows, i, col, header_fmt)
            sheet.write_column(skiprows+1, i, s, border_fmt)

        if not (autofilter or conditional_value):
            return

        rg = (skiprows, 0, len(df), len(df.columns) - 1)

        if autofilter:
            sheet.autofilter(*rg)

        if conditional_value:
            fmt = book.add_format(
                dict(bg_color='#FFC7CE',
                     font_color='#9C0006',
                     ))

            sheet.conditional_format(
                *rg,
                dict(type='text',
                    criteria='containing',
                    value=conditional_value,
                    format= fmt))

def to_border(df, *arg, **prop):
    "prop : CSS style setting dict values"
    kw = {"border-style": "solid", "border-width": "thin", "vertical-align": "center", **prop}
    return df.style.set_properties(*arg, **kw)

def read_json(f, concatenate=True, *args, **kw):
    dfh = _dfhandler("read_json", f, *args, **kw)
    return df_cast(dfh.compute(concatenate))

def df_cast(df):
    """
    smaller cast dataframe (destroied function)
    Parameters
    ----------
    df: pd.DataFrame
    return : inplace dataframe
    """
    if isdataframe(df) and all(np.dtype("O") is x for x in df.dtypes):
        return df

    int_cols = df.select_dtypes(include=['int']).columns.tolist()
    for col in int_cols:
        if ((np.max(df[col]) <= 127) and(np.min(df[col] >= -128))):
            #print('cast {} : {} -> {}'.format(col, df[col].dtype, 'np.int8'))
            df[col] = df[col].astype(np.int8)
        elif ((np.max(df[col]) <= 32767) and(np.min(df[col] >= -32768))):
            #print('cast {} : {} -> {}'.format(col, df[col].dtype, 'np.int16'))
            df[col] = df[col].astype(np.int16)
        elif ((np.max(df[col]) <= 2147483647) and(np.min(df[col] >= -2147483648))):
            df[col] = df[col].astype(np.int32)
        else:
            df[col] = df[col].astype(np.int64)

    # float
    float_cols = df.select_dtypes(include=['float']).columns.tolist()
    for col in float_cols:
        df[col] = df[col].astype(np.float32)

    return df

def dflines(filepath_or_buffer, encoding=None, columns = ["filename", "linenumber", "line", "size"], linestrip=True):
    def dfline(path):
        enc = encoding or path.encoding
        with path.open(mode="rb") as r:
            for i, line in enumerate(r, 1):
                if linestrip:
                    line = line.strip()
                    yield path, i, line.decode(enc), len(line)
                else:
                    yield path, i, line.decode(enc), len(line)

    pathes = list(Path(filepath_or_buffer).lsdir())
    if len(pathes) == 1:
        return pd.DataFrame(dfline(pathes[0]), columns=columns)
    elif len(pathes) > 1:
        return pd.concat(pd.DataFrame(dfline(p), columns=columns) for p in pathes)
    else:
        raise FileNotFoundError("Not Found filepath or buffer `{}`".format(filepath_or_buffer))

def vdf(df, subset):
    exclude = [i for i in subset if i not in subset]
    return [df[subset].reset_index(), df[exclude]]

def hdf(df, cond):
    if not cond.empty and all(cond.dtypes == np.bool):
        return [df[cond], df[~cond]]
    else:
        m = df.loc[cond.index]
        return [m, df.loc[df.index.difference(m.index)]]

def read_any(f, *args, **kw):
    ext = Path(f).ext[1:]

    if ext in ["txt", "tsv", "csv", "zip", "tar.gz", "gz", "bz2", "xz"]:
        return read_csv(f, *args, **kw)

    elif ext.startswith("xls"):
        return read_excel(f, *args, **kw)

    elif ext == "xlb":
        try:
            return read_excel(f, *args, **kw)
        except:
            return read_csv(f, *args, **kw)

    elif ext in ["mdb", "accdb", "db", "sqlite", "sqlite3"]:
        if "dbutil.read_db" not in sys.modules:
            from util.dbutil import read_db
        return read_db(f, *args, **kw)

    elif ext == "json":
        return read_json(f, *args, **kw)

    else:
        raise ValueError("Unknown Format File " + ext)


if __name__ == "__main__":
    """
       TestCase below
    """
    def test():
        from util.core import tdir
        from datetime import datetime as dt

        def nontest_daskwrapper():
            #TODO
            from util.core import zopen
            rf = r"C:\temp\utf8.csv"

            assert(dd.read_csv(Path(rf)).compute().shape == (392, 9))

            with open(rf) as f:
                assert(dd.read_csv(f).compute().shape == (392, 9))

            with open(rf, "rb") as f:
                assert(dd.read_csv(f).compute().shape == (392, 9))

            for f in zopen(r"C:\temp\クエリ1.zip\クエリ1.csv"):
                print(dd.read_csv(f, encoding=f.encoding).compute())

        def test__dfhandler():
            def handler_run(f, concatenate=True):
                ret = _dfhandler("read_csv", f)
                return ret.compute(concatenate)

            handler_run(tdir + "test.csv")
            handler_run(tdir + "test.zip")
            handler_run(tdir + "test.tar.gz")
            handler_run(tdir + "test.tar")
            handler_run(tdir + "test.lzh")
            handler_run(tdir + "test.rar")
            handler_run(tdir + "test.csv.gz")
            handler_run(tdir + "test.csv.xz")
            handler_run(tdir + "test.csv.bz2")
            #handler_run(tdir + "test.csv.tar.gz") #TODO
            handler_run(tdir + "test.tar.gz/test.*")
            handler_run(tdir + "test.tar/test.*")
            handler_run(tdir + "test.lzh/test.*")
            handler_run(tdir + "test.rar/test.*")
            handler_run(tdir + "test.csv.gz/test.*")
            handler_run(tdir + "test.csv.xz/test.*")
            handler_run(tdir + "test.csv.bz2/test.*")

            handler_run(tdir + "test.csv", False)
            handler_run(tdir + "test.zip", False)
            handler_run(tdir + "test.tar.gz", False)
            handler_run(tdir + "test.tar", False)
            handler_run(tdir + "test.lzh", False)
            handler_run(tdir + "test.rar", False)
            handler_run(tdir + "test.csv.gz", False)
            handler_run(tdir + "test.csv.xz", False)
            handler_run(tdir + "test.csv.bz2", False)
            #handler_run(tdir + "test.csv.tar.gz", False) #TODO
            handler_run(tdir + "test.tar.gz/test.*", False)
            handler_run(tdir + "test.tar/test.*", False)
            handler_run(tdir + "test.lzh/test.*", False)
            handler_run(tdir + "test.rar/test.*", False)
            handler_run(tdir + "test.csv.gz/test.*", False)
            handler_run(tdir + "test.csv.xz/test.*", False)
            handler_run(tdir + "test.csv.bz2/test.*", False)

            handler_run(Path(tdir + "test.csv").geturi())
            handler_run(Path(tdir + "test.zip").geturi())
            handler_run(Path(tdir + "test.tar.gz").geturi())
            handler_run(Path(tdir + "test.tar").geturi())
            handler_run(Path(tdir + "test.lzh").geturi())
            handler_run(Path(tdir + "test.rar").geturi())
            handler_run(Path(tdir + "test.csv.gz").geturi())
            handler_run(Path(tdir + "test.csv.xz").geturi())
            handler_run(Path(tdir + "test.csv.bz2").geturi())
            #handler_run(Path(tdir + "test.csv.tar.gz").geturi()) #TODO
            handler_run(Path(tdir + "test.tar.gz/test.*").geturi())
            handler_run(Path(tdir + "test.tar/test.*").geturi())
            handler_run(Path(tdir + "test.lzh/test.*").geturi())
            handler_run(Path(tdir + "test.rar/test.*").geturi())
            handler_run(Path(tdir + "test.csv.gz/test.*").geturi())
            handler_run(Path(tdir + "test.csv.xz/test.*").geturi())
            handler_run(Path(tdir + "test.csv.bz2/test.*").geturi())

            handler_run(open(tdir + "test.csv"))
            handler_run(open(tdir + "test.zip"))
            handler_run(open(tdir + "test.tar.gz"))
            handler_run(open(tdir + "test.tar"))
            handler_run(open(tdir + "test.lzh"))
            handler_run(open(tdir + "test.rar"))
            handler_run(open(tdir + "test.csv.gz"))
            handler_run(open(tdir + "test.csv.xz"))
            handler_run(open(tdir + "test.csv.bz2"))
            #handler_run(open(tdir + "test.csv.tar.gz")) #TODO

            handler_run(open(tdir + "test.csv", "rb"))
            handler_run(open(tdir + "test.zip", "rb"))
            handler_run(open(tdir + "test.tar.gz", "rb"))
            handler_run(open(tdir + "test.tar", "rb"))
            handler_run(open(tdir + "test.lzh", "rb"))
            handler_run(open(tdir + "test.rar", "rb"))
            handler_run(open(tdir + "test.csv.gz", "rb"))
            handler_run(open(tdir + "test.csv.xz", "rb"))
            handler_run(open(tdir + "test.csv.bz2", "rb"))
            #handler_run(open(tdir + "test.csv.tar.gz", "rb")) #TODO

        def test_guesskw():
            anser = {'encoding': 'cp932', 'quoting': 0, 'doublequote': False, 'delimiter': ',', 'quotechar': '"', 'sep': ','}
            f = tdir + "diff1.csv"
            ret = _dfhandler("read_csv", f)
            assert(ret.gk[0][1] ==  anser)
            f = tdir + "test.zip/test.csv"
            ret = _dfhandler("read_csv", f)
            assert(ret.gk[0][1] == anser)

        def test_read_csv():
            f = tdir + "diff1.csv"
            assert(read_csv(f).shape == (389, 9))

        def test_read_excel():
            f = tdir + "diff1.xlsx"
            assert(read_excel(f, 0).shape == (389, 9))

        def read_json():
            pass

        def test_read_any():
            f = tdir + "diff1.csv"
            assert((read_any(f) == pd.read_csv(f, dtype="object", keep_default_na=False, encoding="cp932")).all().all())

            f = tdir + "diff1.xlsx"
            assert((read_any(f,0) == pd.read_excel(f, dtype= "str", keep_default_na=False)).all().all())

            f = tdir + "test.zip"
            # print(read_csv(f))

        def test_dflines():
            f = Path(tdir+"test.csv")
            assert((dflines(f) == pd.DataFrame([[f, 1, "n,aa", 4], [f, 2, "1,1", 3], [f, 3, "2,あ", 4]],
                  columns=["filename", "linenumber", "line", "size"])).all().all())
            try:
                dflines("hoges/dfsiaj")
            except FileNotFoundError:
                pass
            except:
                raise AssertionError

        for x, func in list(locals().items()):
            if x.startswith("test_") and callable(func):
                t1 = dt.now()
                func()
                t2 = dt.now()
                print("{} : time {}".format(x, t2-t1))


    test()

